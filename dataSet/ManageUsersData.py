import openpyxl as openpyxl


class ManageUsersData:
    URL = "manage-users"
    TABLE_COLUMNS = ["User Name", "First Name", "Last Name", "Email", "Role"]
    NEW_USERS_FIELDS = ["User Name", "First Name", "Last Name", "Email", "Role", "Password", "Retype Password"]
    EDIT_USERS_FIELDS = ["First Name", "Last Name", "Email", "Role"]
    CHANGE_PWD_USERS_FIELDS = ["User Name", "Password", "Retype Password"]
    ADD_USER_BTN = 'Add User'
    CHANGE_PASSWORD_BTN = 'Change Password'
    EDIT_BTN = 'Edit'
    DELETE_BTN = 'Delete'
    ROLES = ['Admin', 'Read only']
    DEFAULT_ADD_USER_DATA = [
        # {NEW_USERS_FIELDS[0]: "DEFAULTUSER", NEW_USERS_FIELDS[1]: "DEFAULTFN", NEW_USERS_FIELDS[2]: "DEFAULTLN",
        #  NEW_USERS_FIELDS[3]: "D@D.COM", NEW_USERS_FIELDS[4]: "Read only", NEW_USERS_FIELDS[5]: "admin",
        #  NEW_USERS_FIELDS[6]: "admin"}
    ]
    DEFAULT_EDIT_USER_DATA = []

    DEFAULT_PASSWORD_DATA = [{NEW_USERS_FIELDS[5]: "123456", NEW_USERS_FIELDS[6]: "123456"},
                             {NEW_USERS_FIELDS[5]: "", NEW_USERS_FIELDS[6]: "123456"},
                             {NEW_USERS_FIELDS[5]: "123456", NEW_USERS_FIELDS[6]: "12356"},
                             {NEW_USERS_FIELDS[5]: "123456", NEW_USERS_FIELDS[6]: ""}]
    DELETE_USER = ["z1", "z2"]
    CHANGE_PASSWORD = []
    EXCEL_PATH = 'C:\\Test Automate\\Automation\\dataSet\\data.xlsx'

    @staticmethod
    def getAddTestData(sheetName):
        book = openpyxl.load_workbook(ManageUsersData.EXCEL_PATH)
        sheet = book[sheetName]
        ManageUsersData.DEFAULT_ADD_USER_DATA = []
        for i in range(2, sheet.max_row + 1):
            Dict = {}
            for j in range(2, sheet.max_column + 1):
                val = sheet.cell(row=i, column=j).value
                if val is None:
                    val = ''
                Dict[sheet.cell(row=1, column=j).value] = val
            ManageUsersData.DEFAULT_ADD_USER_DATA.append(Dict)
        return ManageUsersData.DEFAULT_ADD_USER_DATA

    @staticmethod
    def getEditTestData(sheetName):
        book = openpyxl.load_workbook(ManageUsersData.EXCEL_PATH)
        ManageUsersData.DEFAULT_EDIT_USER_DATA = []
        sheet = book[sheetName]
        print("-------------------------------------------------------------")
        print(book)
        for i in range(2, sheet.max_row + 1):
            Dict = {}
            for j in range(2, sheet.max_column + 1):
                val = sheet.cell(row=i, column=j).value
                if val is None:
                    val = ''
                Dict[sheet.cell(row=1, column=j).value] = str(val)
            ManageUsersData.DEFAULT_EDIT_USER_DATA.append(Dict)
        return ManageUsersData.DEFAULT_EDIT_USER_DATA

    @staticmethod
    def getChangePwdTestData():
        book = openpyxl.load_workbook(ManageUsersData.EXCEL_PATH)
        ManageUsersData.CHANGE_PASSWORD = []
        sheet = book['Sheet5']
        print("-------------------------------------------------------------")
        print(book)
        for i in range(2, sheet.max_row + 1):
            Dict = {}
            for j in range(2, sheet.max_column + 1):
                val = sheet.cell(row=i, column=j).value
                if val is None:
                    val = ''
                Dict[sheet.cell(row=1, column=j).value] = str(val)
            ManageUsersData.CHANGE_PASSWORD.append(Dict)
        return ManageUsersData.CHANGE_PASSWORD
